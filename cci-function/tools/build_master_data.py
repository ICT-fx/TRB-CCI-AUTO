"""Génère la master data propre de la fonction CCI depuis un export de ventes.

Entrée  : un classeur de ventes historiques (feuille `data`) au format de l'export
          ProConcept/Power BI, colonnes :
          N° document | Description courte | Référence principale | Date document |
          Clé 1 | Nom 1 | (Quantité finale)
Sortie  : app/master_data.xlsx avec 2 feuilles dédupliquées :
          - `clients`   : customer_code | nom_client                 (1 ligne / client)
          - `catalogue` : customer_code | sku | designation          (1 ligne / client-SKU)

Le SKU (« Référence principale ») est un code article à 4 chiffres. La désignation
(« Description courte ») est déterminée de façon unique par le SKU. Le catalogue sert
à : résoudre la Clé 1 par le nom client, et valider/corriger le SKU d'une commande en
comparant le nom de produit au catalogue du client.

Usage :
    python3 tools/build_master_data.py [chemin_source.xlsx]
    (défaut : tools/ventes_source.xlsx)
"""

from __future__ import annotations

import os
import re
import sys

from openpyxl import Workbook, load_workbook

_HERE = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_SOURCE = os.path.join(_HERE, "ventes_source.xlsx")
_OUTPUT = os.path.join(_HERE, "..", "app", "master_data.xlsx")

_SOURCE_SHEET = "data"
# Index des colonnes dans la feuille source (0-based).
_COL_DESIGNATION = 1  # Description courte
_COL_SKU = 2          # Référence principale
_COL_CODE = 4         # Clé 1
_COL_NAME = 5         # Nom 1

_SKU_RE = re.compile(r"^\d{4}$")


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def build(source_path: str, output_path: str) -> None:
    wb = load_workbook(source_path, read_only=True, data_only=True)
    if _SOURCE_SHEET not in wb.sheetnames:
        raise SystemExit(f"Feuille {_SOURCE_SHEET!r} absente de {source_path}")
    ws = wb[_SOURCE_SHEET]

    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # en-tête

    clients: dict[str, str] = {}          # code -> nom_client (première orthographe vue)
    catalogue: dict[tuple[str, str], str] = {}  # (code, sku) -> designation
    skipped = 0

    for r in rows:
        r = list(r) + [None] * 6
        code = _clean(r[_COL_CODE])
        name = _clean(r[_COL_NAME])
        sku = _clean(r[_COL_SKU])
        designation = _clean(r[_COL_DESIGNATION])

        if not code or not _SKU_RE.match(sku):
            skipped += 1
            continue

        clients.setdefault(code, name)
        catalogue.setdefault((code, sku), designation)

    out = Workbook()
    ws_clients = out.active
    ws_clients.title = "clients"
    ws_clients.append(["customer_code", "nom_client"])
    for code in sorted(clients):
        ws_clients.append([code, clients[code]])

    ws_cat = out.create_sheet("catalogue")
    ws_cat.append(["customer_code", "sku", "designation"])
    for (code, sku) in sorted(catalogue):
        ws_cat.append([code, sku, catalogue[(code, sku)]])

    out.save(output_path)

    print(f"Source            : {source_path}")
    print(f"Clients           : {len(clients)}")
    print(f"Lignes catalogue  : {len(catalogue)}")
    print(f"Lignes ignorées   : {skipped} (code ou SKU non conforme)")
    print(f"Écrit             : {os.path.normpath(output_path)}")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else _DEFAULT_SOURCE
    build(src, _OUTPUT)
