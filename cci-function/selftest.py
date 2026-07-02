"""Test local du cœur déterministe — SANS appel à l'API Anthropic.

L'appel Claude de résolution est remplacé par une réponse simulée (monkeypatch),
si bien que le mapping résolution + le garde-fou master data + la génération Excel
sont testés de bout en bout contre la VRAIE master data (app/master_data.xlsx),
sans ANTHROPIC_API_KEY.

    cd cci-function
    python3 selftest.py

Produit `selftest.out.xlsx` et affiche l'en-tête + la ligne de données.
"""

from __future__ import annotations

import datetime
import io
import sys

from openpyxl import load_workbook

from app import anthropic_client, resolver
from app.anthropic_client import _format_delivery_date
from app.excel_writer import build_error_report, build_workbook
from app.models import (
    OrderExtraction,
    ProductLine,
    suggested_filename,
    validate_order,
    validate_resolution,
)


def _fake_resolution(customer_code, lines):
    """Fabrique une fonction qui imite anthropic_client.resolve_order."""
    def _call(order, master_context):
        return {
            "customer_code": customer_code,
            "customer_status": "ok" if customer_code else "introuvable",
            "lines": lines,
        }
    return _call


def _expected_first_bday(year, month):
    d = datetime.date(year, month, 1)
    while d.weekday() >= 5:
        d += datetime.timedelta(days=1)
    return d.strftime("%d/%m/%Y")


def _check_dates() -> None:
    # ISO complet -> JJ/MM/AAAA
    assert _format_delivery_date("2027-06-14") == "14/06/2027"
    # Mois + année seuls -> 1er jour OUVRÉ du mois
    assert _format_delivery_date("2026-10") == _expected_first_bday(2026, 10)
    assert _format_delivery_date("2026-08") == _expected_first_bday(2026, 8)
    # Déjà au bon format / non reconnu -> inchangé
    assert _format_delivery_date("15/07/2026") == "15/07/2026"
    assert _format_delivery_date("Sem 11 (Mars 2027)") == "Sem 11 (Mars 2027)"
    print("Dates : formats JJ/MM/AAAA + 1er jour ouvré OK.")


def main() -> int:
    _check_dates()

    # Commande pour un client réel du master data (Dea Lens Project, Clé 1 = 2780008).
    # 1re ligne : nom AMBIGU (plusieurs "Vismed Multi 10" au catalogue) -> à vérifier
    # (jaune). 2e ligne : SKU absent mais correspondance évidente -> corrigé, PAS jaune.
    # La date est déjà au format JJ/MM/AAAA (comme après extraction).
    order = OrderExtraction(
        customer_name="Dea Lens Project",
        partner_reference="PO-2026-00042",
        requested_delivery_date="15/07/2026",
        products=[
            ProductLine(designation="Vismed Multi 10", sku="9999", quantity=100),
            ProductLine(designation="Vismed 20", sku=None, quantity=50),
        ],
        comments="commande de test",
        confidence=0.92,
        is_readable=True,
        quality_note=None,
    )

    assert validate_order(order) == [], "La validation de format ne devrait pas rejeter."

    anthropic_client.resolve_order = _fake_resolution(
        "2780008",
        [
            {"designation": "Vismed Multi 10", "input_sku": "9999", "resolved_sku": "1311", "status": "ambigu"},
            {"designation": "Vismed 20", "input_sku": None, "resolved_sku": "1136", "status": "corrige"},
        ],
    )
    resolution = resolver.resolve(order)
    print(f"Résolution : code={resolution.customer_code!r}, matched={resolution.matched}, statut={resolution.status!r}")

    assert resolution.matched and resolution.customer_code == "2780008"
    assert order.products[0].resolved_sku == "1311" and order.products[0].sku_status == "ambigu"
    assert order.products[1].resolved_sku == "1136" and order.products[1].sku_status == "corrige"
    assert validate_resolution(order, resolution) == [], "Commande valide, ne devrait pas être rejetée."

    xlsx = build_workbook(order, resolution, file_name="commande_test.pdf")
    out_path = "selftest.out.xlsx"
    with open(out_path, "wb") as f:
        f.write(xlsx)
    print(f"Excel généré : {out_path} ({len(xlsx)} octets)")

    wb = load_workbook(out_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    data = [c.value for c in ws[2]]
    print("\nColonnes :")
    for h, v in zip(headers, data):
        print(f"  {h:22} = {v!r}")

    row = dict(zip(headers, data))
    # Colonnes retirées : plus de statut SKU, ni Confiance / Statut / Note qualité.
    for removed in (
        "SKU 1 (statut)", "SKU 2 (statut)", "Confiance", "Statut",
        "Note qualité", "Numéro de TVA", "Valeur 1",
    ):
        assert removed not in headers, f"Colonne {removed!r} aurait dû être absente."

    assert row["Clé 1"] == "2780008", "Clé 1 (code Customer) doit être remplie."
    assert row["Nom du client"] == "Dea Lens Project"
    assert row["Référence partenaire"] == "PO-2026-00042"
    assert row["Date de livraison souhaitée"] == "15/07/2026", "Date au format JJ/MM/AAAA."
    assert row["SKU 1"] == "1311" and row["Quantité 1"] == 100
    assert row["SKU 2"] == "1136" and row["Quantité 2"] == 50
    # Nom du fichier = nom composé « client - JJ-MM-AAAA » + extension.
    attendu = suggested_filename("Dea Lens Project", "15/07/2026", "commande_test.pdf")
    assert attendu == "Dea Lens Project - 15-07-2026.pdf", attendu
    assert row["Nom du fichier"] == attendu, row["Nom du fichier"]

    # Surlignage : SKU 1 (ambigu) en jaune, SKU 2 (corrigé évident) PAS en jaune.
    def _fill_rgb(col_label):
        return ws.cell(row=2, column=headers.index(col_label) + 1).fill.fgColor.rgb
    assert _fill_rgb("SKU 1") == "FFFFFF00", "SKU ambigu doit être surligné jaune."
    assert _fill_rgb("SKU 2") != "FFFFFF00", "SKU corrigé évident ne doit PAS être surligné."

    # --- Rejet A-revoir : client introuvable ------------------------------
    order2 = OrderExtraction(
        customer_name="Client Inexistant SARL",
        partner_reference="PO-X",
        requested_delivery_date="01/08/2026",
        products=[ProductLine(designation="Vismed Multi 10", sku="1311", quantity=10)],
        confidence=0.8, is_readable=True,
    )
    anthropic_client.resolve_order = _fake_resolution(
        None,
        [{"designation": "Vismed Multi 10", "input_sku": "1311", "resolved_sku": None, "status": "inconnu"}],
    )
    res2 = resolver.resolve(order2)
    assert validate_resolution(order2, res2), "Un client introuvable doit être rejeté (A-revoir)."
    print(f"\nRejet attendu (client introuvable) : {validate_resolution(order2, res2)}")

    # --- Rejet A-revoir : produit hors catalogue (garde-fou anti-hallucination) --
    order3 = OrderExtraction(
        customer_name="Dea Lens Project",
        partner_reference="PO-Y",
        requested_delivery_date="02/08/2026",
        products=[ProductLine(designation="Produit fantôme", sku="0000", quantity=5)],
        confidence=0.8, is_readable=True,
    )
    anthropic_client.resolve_order = _fake_resolution(
        "2780008",
        [{"designation": "Produit fantôme", "input_sku": "0000", "resolved_sku": "4242", "status": "ok"}],
    )
    res3 = resolver.resolve(order3)
    assert order3.products[0].sku_status == "inconnu", "SKU hors catalogue -> 'inconnu'."
    assert validate_resolution(order3, res3), "Un produit hors catalogue doit être rejeté (A-revoir)."
    print(f"Rejet attendu (produit hors catalogue) : {validate_resolution(order3, res3)}")

    # --- Reporting A-revoir : Nom du fichier | Date de l'erreur | Observation ---
    err_xlsx = build_error_report([
        {"nom_fichier": "TRB Chemedica - 02-07-2026.pdf", "date": "02/07/2026",
         "raison": "client introuvable dans la master data : X", "note": "scan flou"},
        {"nom_fichier": "Client Y - 03-07-2026.png", "date": "03/07/2026",
         "raison": "produit(s) hors catalogue: Ostenil", "note": ""},
    ])
    ews = load_workbook(io.BytesIO(err_xlsx)).active
    assert [c.value for c in ews[1]] == ["Nom du fichier", "Date de l'erreur", "Observation"]
    r1 = [c.value for c in ews[2]]
    assert r1[0] == "TRB Chemedica - 02-07-2026.pdf" and r1[1] == "02/07/2026"
    assert "client introuvable" in r1[2] and "scan flou" in r1[2], r1[2]
    r2 = [c.value for c in ews[3]]
    assert r2[2] == "produit(s) hors catalogue: Ostenil", r2[2]  # pas de note -> pas de tiret
    print("\nReporting A-revoir : colonnes + observations OK.")

    print("\nOK — tous les contrôles passent.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
