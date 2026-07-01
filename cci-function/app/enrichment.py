"""Master data : clients (nom ↔ code Customer) + catalogue produit par client.

Chargée une seule fois (worker chaud) depuis `master_data.xlsx`, elle-même
générée depuis l'historique des ventes par `tools/build_master_data.py`.
Deux feuilles :
  - `clients`   : customer_code | nom_client
  - `catalogue` : customer_code | sku | designation   (1 ligne par client-SKU)

Sert à la résolution (resolver.py) : retrouver le code Customer (Clé 1) d'un
client et valider/corriger le SKU d'une commande via le catalogue de ce client.
Remplaçable plus tard par SharePoint / l'ERP en ne changeant QUE ce fichier.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

MASTER_DATA_PATH = Path(__file__).with_name("master_data.xlsx")
CLIENTS_SHEET = "clients"
CATALOGUE_SHEET = "catalogue"


def _s(value) -> str:
    return "" if value is None else str(value).strip()


class MasterDataStore:
    """Charge la master data une fois et expose clients + catalogues.

    `context` est le bloc texte (clients + catalogues, groupé par client) fourni
    à Claude lors de la résolution ; il est stable entre documents d'un même lot,
    donc mis en cache côté API (voir resolver.py).
    """

    def __init__(self, path: Path = MASTER_DATA_PATH) -> None:
        self._path = path
        self._clients: list[tuple[str, str]] = []            # [(code, nom), …]
        self._catalogue: dict[str, list[tuple[str, str]]] = {}  # code -> [(sku, designation), …]
        self._name_by_code: dict[str, str] = {}
        self._context = ""
        self._loaded = False
        self._available = False

    # --- chargement ------------------------------------------------------
    def _ensure_loaded(self) -> None:
        if self._loaded:
            return
        self._loaded = True
        if not self._path.exists():
            logger.warning("master_data.xlsx introuvable (%s) — résolution désactivée.", self._path)
            return
        try:
            wb = load_workbook(self._path, data_only=True, read_only=True)
            try:
                self._clients = self._read_clients(wb)
                self._name_by_code = {code: name for code, name in self._clients}
                self._catalogue = self._read_catalogue(wb)
            finally:
                wb.close()
            self._context = self._render_context()
            self._available = bool(self._clients)
            logger.info(
                "Master data chargée : %d clients, %d lignes catalogue.",
                len(self._clients),
                sum(len(v) for v in self._catalogue.values()),
            )
        except Exception:  # pragma: no cover - robustesse au chargement
            logger.exception("Échec du chargement du master data — résolution désactivée.")

    def _read_clients(self, wb) -> list[tuple[str, str]]:
        if CLIENTS_SHEET not in wb.sheetnames:
            raise ValueError(f"Feuille '{CLIENTS_SHEET}' absente du master data.")
        ws = wb[CLIENTS_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or ()
        idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        c_code, c_name = idx.get("customer_code"), idx.get("nom_client")
        if c_code is None or c_name is None:
            raise ValueError("Colonnes 'customer_code' / 'nom_client' absentes de la feuille clients.")
        out: list[tuple[str, str]] = []
        for row in rows:
            if row is None:
                continue
            code = _s(row[c_code] if c_code < len(row) else None)
            name = _s(row[c_name] if c_name < len(row) else None)
            if code:
                out.append((code, name))
        return out

    def _read_catalogue(self, wb) -> dict[str, list[tuple[str, str]]]:
        if CATALOGUE_SHEET not in wb.sheetnames:
            raise ValueError(f"Feuille '{CATALOGUE_SHEET}' absente du master data.")
        ws = wb[CATALOGUE_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None) or ()
        idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        c_code, c_sku, c_des = idx.get("customer_code"), idx.get("sku"), idx.get("designation")
        if c_code is None or c_sku is None:
            raise ValueError("Colonnes 'customer_code' / 'sku' absentes de la feuille catalogue.")
        out: dict[str, list[tuple[str, str]]] = {}
        for row in rows:
            if row is None:
                continue
            code = _s(row[c_code] if c_code < len(row) else None)
            sku = _s(row[c_sku] if c_sku < len(row) else None)
            des = _s(row[c_des] if (c_des is not None and c_des < len(row)) else None)
            if code and sku:
                out.setdefault(code, []).append((sku, des))
        return out

    def _render_context(self) -> str:
        """Bloc master data pour Claude, groupé par client (code, nom, catalogue)."""
        lines: list[str] = ["MASTER DATA — clients et leurs catalogues :", ""]
        for code, name in self._clients:
            lines.append(f"[{code}] {name}")
            for sku, des in self._catalogue.get(code, []):
                lines.append(f"  - {sku} | {des}")
        return "\n".join(lines)

    # --- accès -----------------------------------------------------------
    @property
    def available(self) -> bool:
        self._ensure_loaded()
        return self._available

    @property
    def context(self) -> str:
        self._ensure_loaded()
        return self._context

    def customers(self) -> list[tuple[str, str]]:
        self._ensure_loaded()
        return self._clients

    def name_for(self, code: str | None) -> str | None:
        self._ensure_loaded()
        return self._name_by_code.get(_s(code)) if code else None

    def catalogue_for(self, code: str | None) -> list[tuple[str, str]]:
        self._ensure_loaded()
        return self._catalogue.get(_s(code), [])

    def has_sku(self, code: str | None, sku: str | None) -> bool:
        return any(s == _s(sku) for s, _ in self.catalogue_for(code))


# Instance unique réutilisée entre invocations (worker chaud).
_store: MasterDataStore | None = None


def get_store() -> MasterDataStore:
    global _store
    if _store is None:
        _store = MasterDataStore()
    return _store
