"""Génération de l'Excel CCI avec openpyxl — format forcé, déterministe.

Une ligne par commande : les champs d'en-tête une fois (champs extraits + code
Customer « Clé 1 »), puis les produits étalés en colonnes
(SKU i / Quantité i / SKU i (statut)), puis des colonnes de diagnostic.

Le SKU écrit est le SKU RÉSOLU (correct, issu du catalogue du client). Seuls les
SKU AMBIGUS (statut "ambigu" — l'IA a hésité entre plusieurs produits proches)
sont surlignés jaune, pour revue rapide ; les corrections évidentes ne le sont
pas. Claude n'écrit jamais ce fichier : le code impose mécaniquement les colonnes.
"""

from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import OrderExtraction, Resolution

SHEET_NAME = "CCI"
_HEADER_FILL = "FF15578F"  # bleu TRB
_HEADER_FONT = "FFFFFFFF"
_HIGHLIGHT_FILL = "FFFFFF00"  # jaune : à vérifier (SKU ambigu / Clé 1 vide)

# Détecte une chaîne susceptible d'être interprétée comme formule par Excel.
_FORMULA_PREFIX = re.compile(r"^[=+\-@\t\r]")

# 3 colonnes par produit : SKU / Quantité / statut de résolution du SKU.
_PRODUCT_COLS_PER_ITEM = 3

# Colonnes de diagnostic (en fin de ligne).
_DIAGNOSTIC_COLUMNS = ["Nom du fichier", "Confiance", "Statut", "Note qualité"]


def _safe_cell(value):
    """Neutralise l'injection de formules.

    Les valeurs viennent de documents clients arbitraires : une cellule commençant
    par = + - @ tab ou CR pourrait être évaluée comme formule à l'ouverture. On
    préfixe ces chaînes d'une apostrophe. Les nombres ne sont jamais altérés
    (Excel ne les évalue pas).
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return ""
    text = str(value)
    if _FORMULA_PREFIX.match(text):
        return "'" + text
    return text


def _is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _product_headers(max_products: int) -> list[str]:
    headers: list[str] = []
    for i in range(1, max_products + 1):
        headers += [f"SKU {i}", f"Quantité {i}", f"SKU {i} (statut)"]
    return headers


def _finish_sheet(ws, headers: list[str]) -> None:
    """Style d'en-tête + figeage + largeurs de colonnes."""
    header_font = Font(bold=True, color=_HEADER_FONT)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for idx in range(1, len(headers) + 1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = min(max(len(headers[idx - 1]) + 2, 12), 44)


# --- Colonnes d'en-tête fixes (champs extraits + Clé 1) -------------------
# Variante « order object » (build_workbook, 1 ligne) : getter(order, resolution).
_HEADER_COLUMNS: list[tuple[str, object]] = [
    ("Nom du client", lambda o, r: o.customer_name),
    ("Clé 1", lambda o, r: r.customer_code),  # code Customer résolu
    ("Référence partenaire", lambda o, r: o.partner_reference),
    ("Date de livraison souhaitée", lambda o, r: o.requested_delivery_date),
]

# Variante « record dict » (build_consolidated_workbook, N lignes) : getter(record).
_FIXED_HEADER_COLUMNS: list[tuple[str, object]] = [
    ("Nom du client", lambda r: r.get("customer_name")),
    ("Clé 1", lambda r: r.get("customer_code")),
    ("Référence partenaire", lambda r: r.get("partner_reference")),
    ("Date de livraison souhaitée", lambda r: r.get("requested_delivery_date")),
]

_CLE1_LABEL = "Clé 1"


def _resolution(record: dict) -> dict:
    r = record.get("resolution")
    return r if isinstance(r, dict) else {}


def build_consolidated_workbook(rows: list[dict]) -> bytes:
    """Construit l'Excel CCI consolidé : 1 feuille, 1 ligne par record.

    Colonnes : en-têtes fixes, puis `SKU i / Quantité i / SKU i (statut)` pour
    i = 1..MAX, puis diagnostic. Les SKU ambigus (statut "ambigu") et une Clé 1
    manquante sont surlignés jaune.

    `rows` vide ⇒ classeur valide avec uniquement les en-têtes.
    """
    rows = rows or []

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    max_products = max((len(r.get("products") or []) for r in rows), default=0)
    max_products = max(max_products, 1)

    fixed = _FIXED_HEADER_COLUMNS
    headers: list[str] = [label for label, _ in fixed]
    headers += _product_headers(max_products)
    headers += _DIAGNOSTIC_COLUMNS
    ws.append(headers)

    cle1_col = next((i for i, (label, _) in enumerate(fixed, start=1) if label == _CLE1_LABEL), None)
    yellow = PatternFill("solid", fgColor=_HIGHLIGHT_FILL)

    for record in rows:
        res = _resolution(record)
        values: list = [getter(record) for _, getter in fixed]
        products = record.get("products") or []
        statuses: list[str | None] = []
        for i in range(max_products):
            if i < len(products):
                p = products[i] or {}
                values += [p.get("sku"), p.get("quantity"), p.get("sku_status")]
                statuses.append(p.get("sku_status"))
            else:
                values += [None, None, None]
                statuses.append(None)
        values += [
            record.get("filename"),
            record.get("confidence"),
            res.get("status") or "OK",
            record.get("quality_note"),
        ]

        ws.append([_safe_cell(v) for v in values])
        row_idx = ws.max_row

        # Surlignage : Clé 1 vide (anormal) + chaque SKU ambigu (à vérifier).
        if cle1_col and _is_empty(values[cle1_col - 1]):
            ws.cell(row=row_idx, column=cle1_col).fill = yellow
        for i, status in enumerate(statuses):
            if status == "ambigu":
                sku_col = len(fixed) + i * _PRODUCT_COLS_PER_ITEM + 1
                ws.cell(row=row_idx, column=sku_col).fill = yellow

    _finish_sheet(ws, headers)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_workbook(
    order: OrderExtraction, resolution: Resolution, file_name: str
) -> bytes:
    """Construit le classeur .xlsx (1 ligne) et renvoie ses octets."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    n = len(order.products)
    headers: list[str] = [label for label, _ in _HEADER_COLUMNS]
    headers += _product_headers(n)
    headers += _DIAGNOSTIC_COLUMNS
    ws.append(headers)

    row: list = [getter(order, resolution) for _, getter in _HEADER_COLUMNS]
    for p in order.products:
        row += [p.resolved_sku, p.quantity, p.sku_status]
    row += [
        file_name,
        round(order.confidence, 2),
        resolution.status or ("OK" if order.is_readable else "document peu lisible"),
        order.quality_note,
    ]
    ws.append([_safe_cell(v) for v in row])
    row_idx = ws.max_row

    yellow = PatternFill("solid", fgColor=_HIGHLIGHT_FILL)
    cle1_col = next((i for i, (label, _) in enumerate(_HEADER_COLUMNS, start=1) if label == _CLE1_LABEL), None)
    if cle1_col and _is_empty(row[cle1_col - 1]):
        ws.cell(row=row_idx, column=cle1_col).fill = yellow
    for i, p in enumerate(order.products):
        if p.sku_status == "ambigu":
            sku_col = len(_HEADER_COLUMNS) + i * _PRODUCT_COLS_PER_ITEM + 1
            ws.cell(row=row_idx, column=sku_col).fill = yellow

    _finish_sheet(ws, headers)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
